"""The historical archive: census, fetch, and per-era parse.

341 published ZIPs spanning 2006-06 to 2026-08. This module enumerates them,
downloads the registered-adviser population (ERAs are OUT by ruling R1 — they
file a subset Form ADV with no Item 4 and cannot carry the succession signal
even in principle), and parses each into canonical rows via ``lineage``.

Three things this module refuses to do, each because a bug already taught it:

* **Never order by page position.** Selection and "latest" come from a date
  parsed out of the filename. Trusting link order once selected a 2006 file as
  the newest. [I-9]
* **Never classify a population by substring against the path.** The containing
  directory is literally named ``...-exempt-reporting-advisers``, so a path test
  for "exempt" matches every file. Classification reads the filename. [I-10]
* **Never treat an unparseable name as recent.** A file whose date will not
  parse is enumerated and skipped, not sorted to the front or the back.
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

from . import config, lineage
from .errors import FeedParseError, FetchError
from .fetch import Fetcher

ARCHIVE_INDEX_URL = (
    "https://www.sec.gov/data-research/sec-markets-data/"
    "information-about-registered-investment-advisers-exempt-reporting-advisers"
)

_ZIP_HREF_RE = re.compile(r'href="(/files/[^"]+\.zip)"', re.I)

#: Date digits from the FILENAME, with every observed variation tolerated:
#: an optional ``ia`` prefix (``010118-exempt.zip`` has none), 6- or 8-digit
#: dates, an optional revision or population suffix.
_DATE_RE = re.compile(r"^(?:ia)?0?(\d{6}|\d{8})(?:[-_].*)?\.zip$", re.I)


@dataclass(frozen=True)
class ArchiveFile:
    path: str
    filename: str
    family: str
    year: int
    month: int
    day: int
    exempt: bool

    @property
    def url(self) -> str:
        return f"https://www.sec.gov{self.path}"

    @property
    def snapshot_date(self) -> str:
        return f"{self.year:04d}-{self.month:02d}-{self.day:02d}"

    @property
    def month_key(self) -> str:
        return f"{self.year:04d}-{self.month:02d}"


def _parse_date(filename: str) -> tuple[int, int, int] | None:
    m = _DATE_RE.match(filename)
    if not m:
        return None
    d = m.group(1)
    if len(d) == 8:
        mm, dd, yy = int(d[:2]), int(d[2:4]), int(d[4:])
    else:
        mm, dd, yy = int(d[:2]), int(d[2:4]), 2000 + int(d[4:])
    if not (1 <= mm <= 12 and 1 <= dd <= 31 and 2000 <= yy <= 2030):
        return None
    return yy, mm, dd


def _is_exempt(filename: str) -> bool:
    return "exempt" in filename.lower()


def census(fetcher: Fetcher) -> tuple[list[ArchiveFile], list[str]]:
    """Enumerate the archive. Returns (files, undatable_filenames)."""
    html = fetcher.get_text(ARCHIVE_INDEX_URL, surface="archive_index")
    paths = sorted(set(_ZIP_HREF_RE.findall(html)))
    if not paths:
        raise FeedParseError(
            f"no .zip links at {ARCHIVE_INDEX_URL} ({len(html)} bytes) -- page layout may have changed"
        )
    files: list[ArchiveFile] = []
    undatable: list[str] = []
    for p in paths:
        fn = p.rsplit("/", 1)[-1]
        fam = p.rsplit("/", 2)[-2] if p.count("/") >= 2 else "?"
        d = _parse_date(fn)
        if d is None:
            undatable.append(fn)
            continue
        files.append(ArchiveFile(p, fn, fam, d[0], d[1], d[2], _is_exempt(fn)))
    files.sort(key=lambda f: (f.year, f.month, f.day, f.exempt))
    return files, undatable


def registered_only(files: list[ArchiveFile]) -> list[ArchiveFile]:
    """Registered-adviser population. ERAs are OUT by ruling R1."""
    return [f for f in files if not f.exempt]


def coverage_gaps(files: list[ArchiveFile]) -> list[str]:
    """Months with no snapshot, between the first and last observed.

    A gap is a recorded finding, not a silent absence -- Phase B ingests them as
    holes so a diff across one cannot be mistaken for a one-month transition.
    """
    have = {f.month_key for f in files}
    if not have:
        return []
    keys = sorted(have)
    y0, m0 = (int(x) for x in keys[0].split("-"))
    y1, m1 = (int(x) for x in keys[-1].split("-"))
    missing, y, m = [], y0, m0
    while (y, m) <= (y1, m1):
        k = f"{y:04d}-{m:02d}"
        if k not in have:
            missing.append(k)
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    return missing


# --------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------


def _rows_from_zip(payload: bytes) -> tuple[list[str], list[list]]:
    """Return (header, rows) from whichever format this era used."""
    try:
        z = zipfile.ZipFile(io.BytesIO(payload))
    except zipfile.BadZipFile as exc:
        raise FeedParseError(f"payload is not a zip ({len(payload)} bytes)") from exc
    names = [n for n in z.namelist() if not n.endswith("/")]
    if not names:
        raise FeedParseError("zip is empty")
    name = names[0]
    data = z.read(name)

    if name.lower().endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - dependency guard
            raise FeedParseError(
                "openpyxl is required to read XLSX-era snapshots; roughly half the "
                "archive by file count is Excel"
            ) from exc
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [[("" if c is None else str(c)) for c in r] for r in it]
        wb.close()
        return header, rows

    text = data.decode("latin-1")
    sample = text[:8000]
    delim = "|" if sample.count("|") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    header = next(reader, None)
    if header is None:
        raise FeedParseError(f"{name} has no header row")
    return header, [r for r in reader]


@dataclass
class ParsedSnapshot:
    snapshot_date: str
    source_file: str
    era: str
    n_columns: int
    absent_fields: tuple[str, ...]
    rows: dict[str, dict]
    skipped: int

    @property
    def n_rows(self) -> int:
        return len(self.rows)


def parse_snapshot(af: ArchiveFile, payload: bytes) -> ParsedSnapshot:
    """Parse one archive ZIP into canonical CRD-keyed rows."""
    header, raw_rows = _rows_from_zip(payload)
    res = lineage.resolve(header)
    if lineage.JOIN_KEY not in res.index:
        raise FeedParseError(
            f"{af.filename}: no CRD column among {res.n_columns} columns; cannot join this era"
        )
    out: dict[str, dict] = {}
    skipped = 0
    for row in raw_rows:
        crd = res.get(row, lineage.JOIN_KEY)
        if crd:
            crd = crd.replace(".0", "").strip()
        if not crd or not crd.isdigit():
            skipped += 1
            continue
        out[crd] = {
            "crd": crd,
            "legal_name": res.get(row, "legal_name"),
            "sec_status": res.get(row, "sec_status"),
            "filing_date": res.get(row, "filing_date"),
            "total_employees": lineage.normalize_int(res.get(row, "total_employees")),
            "aum_total": lineage.normalize_int(res.get(row, "aum_total")),
            "acquired_name": res.get(row, "acquired_name"),
            "acquired_crd": res.get(row, "acquired_crd"),
        }
    if not out:
        raise FeedParseError(
            f"{af.filename}: parsed to zero CRD-keyed rows from {len(raw_rows)} raw "
            f"-- refusing to treat as empty-ok"
        )
    return ParsedSnapshot(
        snapshot_date=af.snapshot_date,
        source_file=af.filename,
        era=lineage.era_label(res),
        n_columns=res.n_columns,
        absent_fields=res.absent,
        rows=out,
        skipped=skipped,
    )


def raw_dir() -> Path:
    d = config.state_home() / "archive"
    d.mkdir(parents=True, exist_ok=True)
    return d


def fetch_archive_file(fetcher: Fetcher, af: ArchiveFile, *, preserve: bool = True) -> bytes:
    """Fetch one archive ZIP, preserving the raw payload by default (B1)."""
    dest = raw_dir() / af.filename
    if dest.exists():
        return dest.read_bytes()
    payload = fetcher.get_bytes(af.url, surface="archive_zip")
    if preserve:
        dest.write_bytes(payload)
    return payload
